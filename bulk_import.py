#!/usr/bin/env python3
"""Bulk import historical audit data from Excel files."""

import json
import sys
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


def import_excel(file_path):
    """Import audit data from Excel file and append to audits.json."""
    data_dir = Path(__file__).parent / 'data'
    data_file = data_dir / 'audits.json'

    # Load existing audits
    if data_file.exists():
        audits = json.loads(data_file.read_text())
    else:
        audits = []

    # Load Excel workbook
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    # Skip header row, process all others
    headers = [cell.value for cell in ws[1]]
    imported_count = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):  # Skip empty rows
            continue

        # Create audit record from row data
        audit = {}
        for col_idx, value in enumerate(row):
            if col_idx < len(headers):
                header = headers[col_idx]
                if header and value is not None:
                    audit[header.lower().replace(' ', '_')] = value

        # Add import timestamp
        audit['timestamp'] = datetime.now().isoformat()
        audit['source'] = 'bulk_import'

        audits.append(audit)
        imported_count += 1

    # Save back
    data_file.write_text(json.dumps(audits, indent=2))
    print(f"✓ Imported {imported_count} audits from {file_path}")
    print(f"Total audits now: {len(audits)}")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python bulk_import.py <excel_file_path> [excel_file_path2 ...]")
        sys.exit(1)

    for file_path in sys.argv[1:]:
        print(f"Importing {file_path}...")
        import_excel(file_path)
