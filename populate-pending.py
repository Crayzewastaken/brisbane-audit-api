#!/usr/bin/env python3
"""Populate audits_pending.json from existing Excel audit files."""

import json
import os
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl")
    import openpyxl


def extract_audits_from_excel(file_path):
    """Extract audit rows from Excel file."""
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        audits = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Skip empty rows
                continue

            # Map Excel columns to audit fields
            audit = {
                'business_name': row[0] if len(row) > 0 else 'Unknown',
                'category': row[1] if len(row) > 1 else 'Uncategorized',
                'phone': str(row[2]) if len(row) > 2 and row[2] else 'N/A',
                'email': str(row[3]) if len(row) > 3 and row[3] else 'N/A',
                'package': str(row[4]) if len(row) > 4 and row[4] else 'Unassigned',

                # Audit findings - use columns if they exist
                'website_status': str(row[5]) if len(row) > 5 and row[5] else 'Not assessed',
                'gbp_status': str(row[6]) if len(row) > 6 and row[6] else 'Not assessed',
                'social_media_status': str(row[7]) if len(row) > 7 and row[7] else 'Not assessed',

                # Business metrics - defaults for now
                'total_cost': 2500,  # Default cost
                'total_profit': 8000,  # Default profit
                'time_to_build_hours': 40,  # Default hours

                # Outreach & conversion - defaults
                'outreach_status': 'not_contacted',
                'securing_probability': 50,  # Middle estimate
                'wants_bundle': False,
                'bundle_details': '',

                # Scoring - defaults
                'worth_score': 3,  # Middle score
                'confidence_level': 3,  # Middle confidence
                'notes': f'Imported from {Path(file_path).name}',

                'timestamp': datetime.now().isoformat()
            }

            audits.append(audit)

        return audits

    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return []


def main():
    # Find audit files
    downloads = Path.home() / 'Downloads'
    audit_files = list(downloads.glob('*audit*.xlsx')) + list(downloads.glob('*Audit*.xlsx'))

    if not audit_files:
        print("No audit Excel files found")
        return

    print(f"Found {len(audit_files)} audit file(s)")

    # Extract audits from all files
    all_audits = []
    for file_path in audit_files:
        print(f"\nReading {file_path.name}...")
        audits = extract_audits_from_excel(str(file_path))
        all_audits.extend(audits)
        print(f"  + {len(audits)} records")

    # Save to pending file
    pending_file = Path.home() / 'audits_pending.json'

    with open(pending_file, 'w') as f:
        json.dump(all_audits, f, indent=2)

    print(f"\n[OK] Saved {len(all_audits)} audits to {pending_file}")
    print(f"\nNow in your routine, call: await syncPendingAudits()")
    print("This will push all audits to the API.")


if __name__ == '__main__':
    main()
